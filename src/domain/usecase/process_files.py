from datetime import datetime

import openpyxl
import pandas as pd


class ProcessXlsxFiles:

    def __init__(self, path):
        self.path = path
        self.wb = openpyxl.load_workbook(self.path)
        self.ws = self.wb["Extracto"]

    def get_row_final_data(self):
        final_founded_boolean = True
        i = 0
        while final_founded_boolean:
            i += 1
            cell = 'B{}'.format(i)
            if self.ws[cell].value == 'FIN ESTADO DE CUENTA':
                final_founded_boolean = False
        return i

    def get_limits_data(self):
        rows = []
        row_final_data = self.get_row_final_data()
        for i in range(1, row_final_data + 1):
            cell = 'A{}'.format(i)
            cell_value = self.ws[cell].value
            if cell_value == 'Movimientos:':
                rows.append(i)
        return rows, row_final_data

    def get_data(self):
        rows, row_final_data = self.get_limits_data()
        path_data = self.path.split("_")
        month, year = path_data[-1], path_data[-2]
        n = len(rows)
        df = pd.DataFrame()
        for i, val in enumerate(rows):
            ceil = val + 1
            floor = row_final_data - 1
            if i < n-1:
                floor = rows[i+1] - 9
            data = [[cell.value for cell in row] for row in self.ws['A{}:F{}'.format(ceil, floor)]]
            temporal_df = pd.DataFrame(data[1:], columns=data[0])
            df = pd.concat([df, temporal_df], axis=0)
            df['VALOR'] = df['VALOR'].astype(str)
            df['VALOR'] = df['VALOR'].apply(lambda x: x.replace(',', '')).astype(float)
        self.calculate_date(df, year)
        return df

    def calculate_date(self, df, year):
        date = df["FECHA"].apply(lambda x: self.get_date(year, x))
        df["FECHA"] = date

    @staticmethod
    def get_date(year, x):
        data = x.split("/")
        return datetime(int(year), int(data[-1]), int(data[-2]))

    def get_date_data(self):
        list_name = self.path.split('_')
        return int(list_name[-1]), int(list_name[-2])



